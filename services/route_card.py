"""Per-route summary page ("rutekort").

Aggregates everything the app knows about a route into a single printable HTML
view: header (length, 3D, ascent, Naismith), route map (Kartverket topo tiles
with the polyline overlaid), elevation profile (inline SVG from cache),
GPS-fasit measured factor, photos near the route, dagbok/inspeksjon/dugnad/
arbeid entries, and the current validation status. No JavaScript; uses
print-friendly CSS so the user can ⌘P → "Save as PDF" for an offline copy.
"""
from __future__ import annotations

import base64
import html
import math
from typing import Any, Dict, List, Optional, Tuple

from psycopg.rows import dict_row


KM_PER_HOUR = 4.828
MIN_PER_M_ASCENT = 0.1


def _fmt_km(m: Optional[float]) -> str:
    if m is None:
        return "–"
    km = m / 1000
    return f"{km:.1f} km" if km < 10 else f"{round(km)} km"


def _fmt_date(iso: Optional[str]) -> str:
    if not iso:
        return ""
    return iso[:10]


def _naismith(distance_m: Optional[float], ascent_m: Optional[float]) -> Optional[str]:
    if distance_m is None or ascent_m is None:
        return None
    total = (distance_m / 1000 / KM_PER_HOUR) * 60 + ascent_m * MIN_PER_M_ASCENT
    h = int(total // 60)
    m = int(round(total % 60))
    return f"{h} t {m} min" if h > 0 else f"{m} min"


def gather(conn, op_conn, area_code: str, rutenummer: str) -> Dict[str, Any]:
    """Pull every piece of data the route card renders, into one dict."""
    out: Dict[str, Any] = {
        "area_code": area_code,
        "rutenummer": rutenummer,
        "rutenavn": None,
        "length_m": None, "length_3d_m": None,
        "ascent_m": None, "descent_m": None, "min_z": None, "max_z": None,
        "samples": None, "datakilde": None,
        "geometry": None,         # GeoJSON (WGS84) for the route shape SVG
        "start_name": None, "end_name": None,
        "measured_factor": None, "assumed_factor": None, "n_tracks_used": 0,
        "photos": [], "diary": [], "inspections": [], "dugnads": [],
        "work_markers": [], "validation": None, "kulturminner": [],
    }

    # rutenavn + geometry via fotruteinfo + route_link_graph
    with conn.cursor() as cur:
        cur.execute(
            "SELECT MAX(rutenavn) FROM ops.fotruteinfo_patched WHERE rutenummer = %s",
            (rutenummer,),
        )
        row = cur.fetchone()
        if row and row[0] and str(row[0]).strip().lower() != "ukjent":
            out["rutenavn"] = row[0]
        cur.execute(
            """
            SELECT ST_Length(ST_Collect(geom)),
                   ST_AsGeoJSON(ST_Transform(ST_CollectionExtract(ST_Collect(geom), 2), 4326))::json
            FROM ops.route_link_graph WHERE rutenummer = %s
            """,
            (rutenummer,),
        )
        row = cur.fetchone()
        if row:
            out["length_m"] = row[0]
            out["geometry"] = row[1]

    # cached elevation profile
    with op_conn.cursor(row_factory=dict_row) as cur:
        cur.execute(
            "SELECT * FROM ops.route_elevation_cache WHERE rutenummer = %s",
            (rutenummer,),
        )
        ev = cur.fetchone()
    if ev:
        out["length_3d_m"] = ev["length_3d_m"]
        out["ascent_m"] = ev["ascent_m"]
        out["descent_m"] = ev["descent_m"]
        out["min_z"] = ev["min_z"]
        out["max_z"] = ev["max_z"]
        out["samples"] = ev["samples"]
        out["datakilde"] = ev["datakilde"]

    # GPS-fasit
    from .gpx_tracks import compare_to_route
    cmp_ = compare_to_route(conn, area_code, rutenummer)
    out["measured_factor"] = cmp_["measured_factor"]
    out["n_tracks_used"] = cmp_["n_tracks_used"]
    from .operational_store import get_distance_correction_factor
    out["assumed_factor"] = get_distance_correction_factor(op_conn, area_code)

    # photos near route
    from . import field_photos as fp_svc
    out["photos"] = fp_svc.list_photos_near_route(conn, area_code, rutenummer)

    # cultural-heritage monuments within 50 m (Riksantikvaren)
    try:
        from .kulturminne_service import get_kulturminner_near_route
        out["kulturminner"] = get_kulturminner_near_route(rutenummer, radius_m=50).get("kulturminner", [])
    except Exception as e:
        print(f"[route_card] kulturminner lookup failed for {rutenummer}: {e}")

    # annotations: split by kind
    from . import route_annotations as ann_svc
    anns = ann_svc.list_for_route(op_conn, area_code, rutenummer)
    for a in anns:
        k = a.get("kind")
        if k == "diary":
            out["diary"].append(a)
        elif k == "inspection":
            out["inspections"].append(a)
        elif k == "dugnad":
            out["dugnads"].append(a)
        elif k and k.startswith("work_"):
            out["work_markers"].append(a)

    # endpoint names (for "X km fra <start>, Y km fra <end>" on work items)
    from .elevation import _resolve_endpoint_names
    out["start_name"], out["end_name"] = _resolve_endpoint_names(conn, rutenummer)

    # Position-along for each work marker that has a point but no stored
    # position. ST_LineLocatePoint only accepts a single LineString, so we
    # ST_LineMerge first and skip if the route is disconnected (MultiLineString
    # after merge). UTM 32V coords always set when lon/lat are present.
    _enrich_work_markers(conn, out)

    # validation
    from .route_corrections import validate_route
    out["validation"] = validate_route(conn, rutenummer)
    return out


def _enrich_work_markers(conn, data: Dict[str, Any]) -> None:
    from .sign_candidates import format_utm32v_block
    rn = data["rutenummer"]
    length_m = data.get("length_m") or 0
    # Pre-merge once for distance-along lookups.
    merged_is_linestring = False
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT ST_GeometryType(ST_LineMerge(ST_Collect(geom)))
              FROM ops.route_link_graph WHERE rutenummer = %s
            """,
            (rn,),
        )
        r = cur.fetchone()
        if r and r[0] == "ST_LineString":
            merged_is_linestring = True
    for m in data["work_markers"]:
        lon, lat = m.get("lon"), m.get("lat")
        if lon is not None and lat is not None:
            m["utm32v"] = format_utm32v_block(lon, lat)
        else:
            m["utm32v"] = None
        m["from_start_m"] = None
        m["from_end_m"] = None
        if m.get("position_along_m") is not None and length_m:
            m["from_start_m"] = float(m["position_along_m"])
            m["from_end_m"] = max(0.0, length_m - m["from_start_m"])
            continue
        if lon is None or lat is None or not merged_is_linestring:
            continue
        with conn.cursor() as cur:
            cur.execute(
                """
                WITH g AS (
                    SELECT ST_LineMerge(ST_Collect(geom)) AS geom
                      FROM ops.route_link_graph WHERE rutenummer = %s
                )
                SELECT ST_LineLocatePoint(
                           g.geom,
                           ST_Transform(ST_SetSRID(ST_MakePoint(%s, %s), 4326), 25833)
                       ) * ST_Length(g.geom)
                  FROM g
                """,
                (rn, lon, lat),
            )
            r = cur.fetchone()
            if r and r[0] is not None:
                m["from_start_m"] = float(r[0])
                m["from_end_m"] = max(0.0, length_m - float(r[0]))


# ---------------------------------------------------------------------------
# SVG helpers — both rendered server-side so the page prints without JS.
# ---------------------------------------------------------------------------

def _flatten_lonlat(geometry: Optional[Dict[str, Any]]) -> List[Tuple[float, float]]:
    if not geometry:
        return []
    t = geometry.get("type")
    coords = geometry.get("coordinates") or []
    if t == "LineString":
        return [(c[0], c[1]) for c in coords if len(c) >= 2]
    if t == "MultiLineString":
        return [(c[0], c[1]) for line in coords for c in line if len(c) >= 2]
    return []


def _bbox(points: List[Tuple[float, float]]) -> Optional[Tuple[float, float, float, float]]:
    if not points:
        return None
    lons = [p[0] for p in points]
    lats = [p[1] for p in points]
    return min(lons), min(lats), max(lons), max(lats)


def _pick_zoom_for_bbox(
    min_lon: float, min_lat: float, max_lon: float, max_lat: float,
    width_px: int, height_px: int,
) -> int:
    """Largest WMTS zoom where the bbox fits inside (width_px × height_px)."""
    from .map_snippet import _lonlat_to_global_pixel
    # Walk zoom levels from high to low; first one that fits wins.
    for z in range(17, 4, -1):
        x_min, y_max = _lonlat_to_global_pixel(min_lon, min_lat, z)
        x_max, y_min = _lonlat_to_global_pixel(max_lon, max_lat, z)
        if (x_max - x_min) <= width_px and (y_max - y_min) <= height_px:
            return z
    return 5


def _png_data_url(png_bytes: bytes, mime: str = "image/jpeg") -> str:
    return f"data:{mime};base64,{base64.b64encode(png_bytes).decode('ascii')}"


def _route_overview_map(
    geometry: Optional[Dict[str, Any]],
    work_markers: Optional[List[Dict[str, Any]]] = None,
    width: int = 540, height: int = 320,
    tile_cache: Optional[Dict] = None,
) -> str:
    """Topo-tile map of the whole route. Returns an <img> tag with data URL,
    or "" if there's no geometry to draw."""
    pts = _flatten_lonlat(geometry)
    bb = _bbox(pts)
    if bb is None:
        return ""
    min_lon, min_lat, max_lon, max_lat = bb
    # Pad bbox 5% so the line doesn't kiss the edge.
    pad_lon = (max_lon - min_lon) * 0.05 or 1e-4
    pad_lat = (max_lat - min_lat) * 0.05 or 1e-4
    bb_p = (min_lon - pad_lon, min_lat - pad_lat, max_lon + pad_lon, max_lat + pad_lat)
    zoom = _pick_zoom_for_bbox(*bb_p, width_px=width, height_px=height)
    center_lon = (bb_p[0] + bb_p[2]) / 2
    center_lat = (bb_p[1] + bb_p[3]) / 2
    extras = [
        (float(m["lon"]), float(m["lat"]), "#e8590c")
        for m in (work_markers or [])
        if m.get("lon") is not None and m.get("lat") is not None and not m.get("resolved_at")
    ]
    from .map_snippet import build_snippet
    img = build_snippet(
        center_lon=center_lon, center_lat=center_lat,
        width_px=width, height_px=height, zoom=zoom,
        route_geoms={"_": geometry} if geometry else None,
        tile_cache=tile_cache,
        draw_center_marker=False,
        extra_markers=extras,
    )
    return (
        f'<img src="{_png_data_url(img)}" alt="kart" '
        f'style="width:100%;max-width:{width}px;height:auto;border:1px solid #dee2e6;border-radius:4px"/>'
    )


def _marker_inset_map(
    lon: float, lat: float,
    geometry: Optional[Dict[str, Any]],
    width: int = 220, height: int = 150,
    zoom: int = 14,
    tile_cache: Optional[Dict] = None,
) -> str:
    """Small topo-tile inset centred on a work marker, with the route line
    drawn underneath and a dot on the marker itself."""
    from .map_snippet import build_snippet
    img = build_snippet(
        center_lon=lon, center_lat=lat,
        width_px=width, height_px=height, zoom=zoom,
        route_geoms={"_": geometry} if geometry else None,
        tile_cache=tile_cache,
        draw_center_marker=True,
    )
    return (
        f'<img src="{_png_data_url(img)}" alt="" '
        f'style="width:{width}px;height:{height}px;border:1px solid #dee2e6;border-radius:4px;'
        f'display:block;margin-top:6px"/>'
    )


def _elevation_svg(samples: Optional[List[List[Optional[float]]]]) -> str:
    if not samples:
        return ""
    pts = [(s[0], s[1]) for s in samples if s[1] is not None]
    if len(pts) < 2:
        return ""
    W, H, PAD = 440, 110, 8
    xs = [p[0] for p in pts]
    zs = [p[1] for p in pts]
    min_x, max_x = min(xs), max(xs)
    min_z, max_z = min(zs), max(zs)
    span_x = max_x - min_x or 1
    span_z = max_z - min_z or 1

    def sx(x):
        return PAD + (x - min_x) / span_x * (W - 2 * PAD)

    def sy(z):
        return PAD + (1 - (z - min_z) / span_z) * (H - 2 * PAD)

    line = " ".join(
        f"{'M' if i == 0 else 'L'}{sx(x):.1f},{sy(z):.1f}" for i, (x, z) in enumerate(pts)
    )
    area = f"{line} L{sx(max_x):.1f},{H - PAD:.1f} L{sx(min_x):.1f},{H - PAD:.1f} Z"
    return (
        f'<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" '
        f'style="width:100%;max-width:{W}px;height:auto;background:#f8f9fa;'
        f'border:1px solid #dee2e6;border-radius:4px">'
        f'<path d="{area}" fill="#a5d8ff" opacity="0.6"/>'
        f'<path d="{line}" fill="none" stroke="#1971c2" stroke-width="1.2"/>'
        f'</svg>'
    )


# ---------------------------------------------------------------------------
# Page assembly
# ---------------------------------------------------------------------------

_CSS = """
* { box-sizing: border-box; }
body { font: 14px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
       color: #212529; margin: 0; padding: 24px; max-width: 900px; }
h1 { margin: 0 0 4px; font-size: 22px; }
h2 { font-size: 14px; text-transform: uppercase; letter-spacing: 0.04em;
     color: #495057; border-bottom: 1px solid #dee2e6; padding-bottom: 4px;
     margin: 28px 0 10px; }
.muted { color: #6c757d; font-size: 12px; }
.stats { display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px 16px;
         margin: 10px 0 20px; }
.stat-label { font-size: 10px; text-transform: uppercase; color: #6c757d; }
.stat-value { font-size: 16px; font-weight: 600; }
.row { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; align-items: start; }
.entry { border-left: 3px solid #dee2e6; padding: 4px 10px; margin: 6px 0; }
.entry .meta { font-size: 11px; color: #6c757d; }
.entry .body { white-space: pre-wrap; margin-top: 2px; }
.work-open { border-left-color: #e8590c; }
.thumbs { display: grid; grid-template-columns: repeat(auto-fill, minmax(120px,1fr));
          gap: 4px; margin: 8px 0; }
.thumbs img { width: 100%; aspect-ratio: 1/1; object-fit: cover; border-radius: 3px; }
.badge { display: inline-block; padding: 1px 6px; border-radius: 3px;
         font-size: 10px; font-weight: 600; }
.badge-err  { background:#ffe0e0; color:#c92a2a; }
.badge-warn { background:#fff4d6; color:#a07106; }
.badge-ok   { background:#e5f5e0; color:#2b8a3e; }
.dl { color: #1971c2; text-decoration: none; }
@media print {
  body { max-width: none; padding: 0; }
  h2 { break-after: avoid; }
  .entry { break-inside: avoid; }
  .no-print { display: none !important; }
}
"""


def _esc(s: Optional[str]) -> str:
    return html.escape(s) if s else ""


def _entry_html(
    a: Dict[str, Any],
    work: bool = False,
    *,
    geometry: Optional[Dict[str, Any]] = None,
    start_name: Optional[str] = None,
    end_name: Optional[str] = None,
    tile_cache: Optional[Dict] = None,
) -> str:
    klass = "entry work-open" if work and not a.get("resolved_at") else "entry"
    title = a.get("title") or "(uten tittel)"
    meta_bits = []
    if a.get("occurred_at"):
        meta_bits.append(_fmt_date(a["occurred_at"]))
    if a.get("recorded_by"):
        meta_bits.append(_esc(a["recorded_by"]))
    if work and a.get("kind"):
        meta_bits.append(_esc(a["kind"].replace("work_", "")))
    if work and a.get("resolved_at"):
        meta_bits.append("løst " + _fmt_date(a["resolved_at"]))
    meta = " · ".join(meta_bits)
    body = _esc(a.get("body")) if a.get("body") else ""

    # Per-work-item enrichment: "X km fra <start>, Y km fra <end>", UTM 32V,
    # and a topo inset around the marker.
    pos_html = ""
    inset_html = ""
    if work:
        pos_bits: List[str] = []
        if a.get("from_start_m") is not None:
            pos_bits.append(f"{a['from_start_m'] / 1000:.1f} km fra {_esc(start_name) or 'start'}")
        if a.get("from_end_m") is not None:
            pos_bits.append(f"{a['from_end_m'] / 1000:.1f} km fra {_esc(end_name) or 'ende'}")
        if a.get("utm32v"):
            pos_bits.append(f"UTM 32V {_esc(a['utm32v'])}")
        if pos_bits:
            pos_html = f'<div class="meta">{" · ".join(pos_bits)}</div>'
        if a.get("lon") is not None and a.get("lat") is not None and geometry:
            try:
                inset_html = _marker_inset_map(
                    float(a["lon"]), float(a["lat"]), geometry,
                    tile_cache=tile_cache,
                )
            except Exception as e:
                print(f"[route_card] inset map failed for annotation {a.get('id')}: {e}")
    return (
        f'<div class="{klass}"><div><strong>{_esc(title)}</strong></div>'
        f'<div class="meta">{meta}</div>'
        + pos_html
        + (f'<div class="body">{body}</div>' if body else "")
        + inset_html
        + "</div>"
    )


def render_html(data: Dict[str, Any]) -> str:
    rn = data["rutenummer"]
    navn = data.get("rutenavn") or rn
    length_m = data["length_m"]
    length_3d_m = data["length_3d_m"]
    ascent_m = data["ascent_m"]
    descent_m = data["descent_m"]
    naismith = _naismith(length_m, ascent_m)
    measured = data["measured_factor"]
    assumed = data["assumed_factor"]

    # stats grid
    stats = [
        ("Lengde (2D)", _fmt_km(length_m)),
        ("Lengde (3D)", _fmt_km(length_3d_m) if length_3d_m else "–"),
        ("Stigning", f"{round(ascent_m)} m" if ascent_m else "–"),
        ("Tid (Naismith)", naismith or "–"),
        ("Høyde min–max", f"{round(data['min_z'])}–{round(data['max_z'])} m"
            if data.get("min_z") is not None and data.get("max_z") is not None else "–"),
        ("Fall", f"{round(descent_m)} m" if descent_m else "–"),
        ("Målt faktor", f"{measured}× (av {data['n_tracks_used']} spor)"
            if measured is not None else "–"),
        ("Antatt faktor", f"{assumed}×"),
    ]
    stats_html = "".join(
        f'<div><div class="stat-label">{lbl}</div><div class="stat-value">{val}</div></div>'
        for lbl, val in stats
    )

    # validation badge
    vstatus = (data.get("validation") or {}).get("status", "OK")
    vbadge_cls = {"ERROR": "badge-err", "WARNING": "badge-warn", "OK": "badge-ok"}.get(vstatus, "badge-ok")
    vissues = (data.get("validation") or {}).get("errors", []) + (data.get("validation") or {}).get("warnings", [])
    val_lines = "".join(
        f'<div class="muted">• {_esc(i.get("type",""))}: {_esc((i.get("message") or "")[:140])}</div>'
        for i in vissues[:10]
    ) or '<div class="muted">Ingen feil eller advarsler.</div>'

    # sections
    diary = "".join(_entry_html(a) for a in data["diary"]) or '<div class="muted">Ingen dagboknotater.</div>'
    inspections = "".join(_entry_html(a) for a in data["inspections"]) or '<div class="muted">Ingen inspeksjoner.</div>'
    dugnads = "".join(_entry_html(a) for a in data["dugnads"]) or '<div class="muted">Ingen dugnader.</div>'
    work_open = [a for a in data["work_markers"] if not a.get("resolved_at")]
    work_closed = [a for a in data["work_markers"] if a.get("resolved_at")]
    # Tiles are reused between overview + per-item insets at the same zoom, and
    # adjacent items share a lot of tiles, so caching collapses the wall-time
    # cost of many work items dramatically.
    tile_cache: Dict = {}
    geometry = data.get("geometry")
    start_name = data.get("start_name")
    end_name = data.get("end_name")
    work_kwargs = dict(geometry=geometry, start_name=start_name, end_name=end_name, tile_cache=tile_cache)
    work_html = (
        "".join(_entry_html(a, work=True, **work_kwargs) for a in work_open)
        + ('<div class="muted" style="margin-top:8px">Løste:</div>' if work_closed else "")
        + "".join(_entry_html(a, work=True, **work_kwargs) for a in work_closed[:10])
    ) or '<div class="muted">Ingen arbeidsbehov.</div>'
    try:
        overview_map_html = _route_overview_map(geometry, work_open, tile_cache=tile_cache)
    except Exception as e:
        print(f"[route_card] overview map failed: {e}")
        overview_map_html = ""

    km_items = data.get("kulturminner") or []
    if km_items:
        rows = []
        for k in km_items:
            navn = _esc(k.get("navn") or "Uten navn")
            dist = f' · {k["distance_m"]} m' if k.get("distance_m") is not None else ""
            kat = f' · {_esc(str(k["kategori"]))}' if k.get("kategori") else ""
            link = f' · <a href="{_esc(str(k["link"]))}" target="_blank" rel="noopener">Kulturminnesøk ↗</a>' if k.get("link") else ""
            rows.append(f'<div class="entry"><strong>{navn}</strong>{dist}{kat}{link}</div>')
        kulturminner_html = (
            f'<div class="muted">{len(km_items)} kulturminne(r) innen 50 m — vis hensyn ved arbeid.</div>'
            + "".join(rows)
        )
    else:
        kulturminner_html = '<div class="muted">Ingen kulturminner innen 50 m.</div>'

    photos = data["photos"][:8]
    photos_html = (
        '<div class="thumbs">'
        + "".join(f'<img src="{_esc(p["thumb_url"])}" alt=""/>' for p in photos)
        + "</div>"
        if photos else '<div class="muted">Ingen bilder i nærheten av ruta.</div>'
    )

    area = data["area_code"]
    dagbok_url = f"/api/v1/routes/{_esc(area)}/{_esc(rn)}/dagbok.xlsx"

    return f"""<!doctype html>
<html lang="nb"><head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{_esc(rn)} — {_esc(navn)} · rutekort</title>
<style>{_CSS}</style>
</head><body>

<h1>{_esc(navn)}</h1>
<div class="muted">{_esc(rn)} · {_esc(area)}
  · <span class="badge {vbadge_cls}">{vstatus}</span>
  · <a class="dl no-print" href="{dagbok_url}">Last ned dagbok (Excel)</a>
</div>

<div class="stats">{stats_html}</div>

<div class="row">
  <div>
    <h2>Kart</h2>
    {overview_map_html}
  </div>
  <div>
    <h2>Høydeprofil</h2>
    {_elevation_svg(data.get("samples"))}
  </div>
</div>

<h2>Validering</h2>
{val_lines}

<h2>Kulturminner</h2>
{kulturminner_html}

<h2>Bilder</h2>
{photos_html}

<h2>Dagbok</h2>
{diary}

<h2>Inspeksjoner</h2>
{inspections}

<h2>Dugnader</h2>
{dugnads}

<h2>Arbeidsbehov</h2>
{work_html}

</body></html>
"""


def build_dagbok_xlsx(diary: List[Dict[str, Any]], area_code: str, rutenummer: str) -> bytes:
    """Diary entries as a one-sheet Excel (download from the route card page)."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Dagbok"
    headers = [("Dato", 14), ("Tittel", 30), ("Notat", 60), ("Registrert av", 24)]
    for i, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A3A5C")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    r = 2
    for a in sorted(diary, key=lambda a: a.get("occurred_at") or "", reverse=True):
        ws.cell(row=r, column=1, value=_fmt_date(a.get("occurred_at")))
        ws.cell(row=r, column=2, value=a.get("title") or "")
        c = ws.cell(row=r, column=3, value=a.get("body") or "")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4, value=a.get("recorded_by") or "")
        r += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
