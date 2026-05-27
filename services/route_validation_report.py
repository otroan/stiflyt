"""Per-area route-validation XLSX report (signs_app ExportTab download).

Runs the validators registered in services.validators on every route in the
area, plus a ferry-/boat-track heuristic (sparse-vertex fotruter — see
scripts/detect_boat_segments.py). Output is a two-sheet workbook:

- "Funn"        — one row per issue, with a suggested Kartverket action.
- "Sammendrag"  — per-route status + counts by severity.

Everything goes through ops.fotruteinfo_patched so errata-applied rutenummer
remaps are reflected, matching what the signs_app already shows on the map.
"""
from __future__ import annotations

import os
from collections import defaultdict
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from psycopg.rows import dict_row

from .area_routes import area_route_filter_sql
from .database import quote_identifier, validate_schema_name
from .validators import (
    ValidationIssue,
    ValidationResult,
    get_validator_registry,
)

FOTRUTEINFO_VIEW = "ops.fotruteinfo_patched"

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")

SEVERITY_FILL = {
    "ERROR": PatternFill("solid", fgColor="FFE0E0"),
    "WARNING": PatternFill("solid", fgColor="FFF4D6"),
    "INFO": PatternFill("solid", fgColor="E0F2FF"),
    "OK": PatternFill("solid", fgColor="E5F5E0"),
}

FERRY_MIN_LENGTH_M = 1500
FERRY_MIN_M_PER_VERTEX = 150
FERRY_MAX_VERTS = 30

ACTION_HINT_NB = {
    "INCONSISTENT_RUTENAVN": "Sett samme rutenavn på alle segmenter.",
    "INCONSISTENT_RUTETYPE": "Sett samme rutetype på alle segmenter.",
    "INCONSISTENT_VEDLIKEHOLDSANSVARLIG": "Sjekk om ulike vedlikeholdsansvarlige er korrekt — ofte feilregistrering.",
    "INCONSISTENT_GRADERING": "Sett samme gradering på alle segmenter.",
    "RUTENAVN_UKJENT": "Sett et reelt rutenavn (se forslag fra rutenavn-foreslagsvalidatoren).",
    "MISSING_REQUIRED_FIELDS": "Fyll inn manglende rutenummer.",
    "MISSING_RUTENAVN": "Sett et rutenavn.",
    "MISSING_RUTENAVN_SOME_SEGMENTS": "Fyll inn rutenavn på segmentene som mangler.",
    "MISSING_VEDLIKEHOLDSANSVARLIG": "Sett vedlikeholdsansvarlig på ruta.",
    "MISSING_VEDLIKEHOLDSANSVARLIG_SOME_SEGMENTS": "Fyll inn vedlikeholdsansvarlig på segmentene som mangler.",
    "DUPLICATE_RUTENUMMER_IN_SEGMENT": "Slett duplikat fotruteinfo-rad for rutenummeret.",
    "DUPLICATE_RUTENAVN_IN_SEGMENT": "Slett duplikat fotruteinfo-rad for rutenavnet.",
    "DUPLICATE_RUTETYPE_IN_SEGMENT": "Slett duplikat fotruteinfo-rad for rutetypen.",
    "DUPLICATE_GRADERING_IN_SEGMENT": "Slett duplikat fotruteinfo-rad for graderingen.",
    "DUPLICATE_VEDLIKEHOLDSANSVARLIG_IN_SEGMENT": "Slett duplikat fotruteinfo-rad for vedlikeholdsansvarlig.",
    "FERRY_SUSPECT": "Marker som båtrute/fjordkryssing eller fjern feilregistrert linjestrekk.",
    "ROUTE_HAS_LOOP": "Velg hvilken arm som er den reelle ruta og ekskluder den/de andre i signs_app, eller del opp i egne rutenummer.",
    "ROUTE_DISCONNECTED": "Koble sammen de adskilte delene (små brudd er digitaliseringsfeil) eller del opp i egne rutenummer. Avstander kan ikke beregnes før ruta er sammenhengende.",
    "RUTENAVN_SUGGESTION": "(forslag — ikke en feil)",
}


def _route_schema() -> Optional[str]:
    schema = os.getenv("ROUTE_SCHEMA", "stiflyt")
    return schema if validate_schema_name(schema) else None


def _list_rutenummers(conn, area_code: str) -> List[str]:
    match_sql, match_params = area_route_filter_sql(area_code, "rutenummer")
    sql = f"""
        SELECT DISTINCT rutenummer
        FROM ops.route_link_graph
        WHERE {match_sql}
        ORDER BY rutenummer
    """
    with conn.cursor() as cur:
        cur.execute(sql, match_params)
        return [r[0] for r in cur.fetchall()]


def _load_segments_dict(conn, rutenummer: str) -> Dict[str, List[Dict[str, Any]]]:
    """{segment_objid: [fotruteinfo_rows]} for one route.

    Only includes fotruteinfo rows belonging to this rutenummer. Segments
    shared with other routes show up once per other-route registration, but
    those rows are filtered out — having different rutenummer on a shared
    segment is normal in turrutebasen and must not be flagged by the
    metadata validators as an inconsistency.
    """
    sql = f"""
        SELECT
            fi.fotrute_fk::text  AS segment_objid,
            fi.objid             AS fotruteinfo_objid,
            fi.rutenummer,
            fi.rutenavn,
            fi.vedlikeholdsansvarlig,
            fi.rutetype,
            fi.gradering
        FROM {FOTRUTEINFO_VIEW} fi
        WHERE fi.rutenummer = %s
    """
    out: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute(sql, (rutenummer,))
        for r in cur.fetchall():
            out[r["segment_objid"]].append(r)
    return dict(out)


def _detect_ferry_suspects(conn, area_code: str) -> List[Dict[str, Any]]:
    """Sparse-vertex fotruter (likely ferry crossings) within the area.

    Mirrors the heuristic in scripts/detect_boat_segments.py but joined to
    fotruteinfo_patched so only segments belonging to the requested area are
    returned.
    """
    schema = _route_schema()
    if not schema:
        return []
    schema_quoted = quote_identifier(schema)
    match_sql, match_params = area_route_filter_sql(area_code, "fi.rutenummer")
    sql = f"""
        WITH suspects AS (
            SELECT
                f.objid                                        AS fotrute_fk,
                f.lokalid::text                                AS lokalid,
                ROUND(ST_Length(f.senterlinje)::numeric, 0)    AS len_m,
                ST_NPoints(f.senterlinje)                      AS verts,
                ROUND(
                    (ST_Length(f.senterlinje) /
                     GREATEST(ST_NPoints(f.senterlinje) - 1, 1))::numeric,
                    0
                )                                              AS m_per_vertex
            FROM {schema_quoted}.fotrute f
            WHERE ST_Length(f.senterlinje) > %s
              AND ST_NPoints(f.senterlinje) < %s
              AND ST_Length(f.senterlinje) /
                  GREATEST(ST_NPoints(f.senterlinje) - 1, 1) > %s
        ),
        in_area AS (
            SELECT s.*,
                   (SELECT string_agg(DISTINCT fi.rutenummer, ',' ORDER BY fi.rutenummer)
                      FROM {FOTRUTEINFO_VIEW} fi
                     WHERE fi.fotrute_fk = s.fotrute_fk
                       AND {match_sql}) AS rutenummers
              FROM suspects s
        )
        SELECT *
          FROM in_area
         WHERE rutenummers IS NOT NULL
         ORDER BY len_m DESC
    """
    params: List[Any] = [
        FERRY_MIN_LENGTH_M, FERRY_MAX_VERTS, FERRY_MIN_M_PER_VERTEX,
        *match_params,
    ]
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def _route_meta(segments_dict: Dict[str, List[Dict[str, Any]]]) -> Tuple[Optional[str], Optional[str]]:
    rutenavn = None
    vedl = None
    for rows in segments_dict.values():
        for r in rows:
            if not rutenavn and r.get("rutenavn"):
                rutenavn = r["rutenavn"]
            if not vedl and r.get("vedlikeholdsansvarlig"):
                vedl = r["vedlikeholdsansvarlig"]
            if rutenavn and vedl:
                return rutenavn, vedl
    return rutenavn, vedl


def _write_header(ws, columns: List[Tuple[str, str, int]]) -> None:
    for i, (_k, header, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def build_validation_workbook(conn, area_code: str) -> bytes:
    """Build the validation XLSX for one area and return its bytes."""
    registry = get_validator_registry()
    rutenummers = _list_rutenummers(conn, area_code)

    results: List[ValidationResult] = []
    route_meta_by_rn: Dict[str, Tuple[Optional[str], Optional[str]]] = {}
    for rn in rutenummers:
        segments_dict = _load_segments_dict(conn, rn)
        route_meta_by_rn[rn] = _route_meta(segments_dict)
        result = registry.run_validators(
            {"rutenummer": rn, "segments_dict": segments_dict},
            conn,
        )
        results.append(result)

    ferry_suspects = _detect_ferry_suspects(conn, area_code)

    wb = Workbook()

    ws = wb.active
    ws.title = "Funn"
    issue_columns = [
        ("rutenummer", "Rutenummer", 14),
        ("rutenavn", "Rutenavn", 28),
        ("vedlikeholdsansvarlig", "Vedlikeholdsansvarlig", 28),
        ("severity", "Alvorlighet", 13),
        ("type", "Type", 32),
        ("message", "Beskrivelse", 60),
        ("affected_segments", "Berørte segmenter (lokalid/objid)", 36),
        ("action_hint", "Forslag til Kartverket", 50),
    ]
    _write_header(ws, issue_columns)

    row = 2

    def _write_issue_row(rn: str, issue: ValidationIssue) -> None:
        nonlocal row
        sev = issue.severity.value.upper()
        rutenavn, vedl = route_meta_by_rn.get(rn, (None, None))
        action = ACTION_HINT_NB.get(issue.type, "")
        if issue.type == "RUTENAVN_SUGGESTION":
            suggested = issue.metadata.get("suggested_rutenavn")
            if suggested:
                action = f"Forslag: «{suggested}»"
        values = [
            rn,
            rutenavn or "",
            vedl or "",
            sev,
            issue.type,
            issue.message,
            ", ".join(issue.affected_segments or []),
            action,
        ]
        fill = SEVERITY_FILL.get(sev)
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            if fill:
                cell.fill = fill
        row += 1

    for result in results:
        for issue in result.errors:
            _write_issue_row(result.rutenummer, issue)
        for issue in result.warnings:
            _write_issue_row(result.rutenummer, issue)
        for issue in result.info:
            _write_issue_row(result.rutenummer, issue)

    for fs in ferry_suspects:
        rns = [r.strip() for r in (fs.get("rutenummers") or "").split(",") if r.strip()]
        rn_for_meta = rns[0] if rns else ""
        rutenavn, vedl = route_meta_by_rn.get(rn_for_meta, (None, None))
        msg = (
            f"Mistenkt båtrute/fjordkryssing: {int(fs['len_m']):,} m fordelt på "
            f"{fs['verts']} punkter ({int(fs['m_per_vertex'])} m/punkt). "
            "Reelle stier sampler typisk 10–30 m/punkt."
        )
        values = [
            ", ".join(rns),
            rutenavn or "",
            vedl or "",
            "INFO",
            "FERRY_SUSPECT",
            msg,
            fs.get("lokalid") or str(fs.get("fotrute_fk") or ""),
            ACTION_HINT_NB["FERRY_SUSPECT"],
        ]
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.fill = SEVERITY_FILL["INFO"]
        row += 1

    ws2 = wb.create_sheet("Sammendrag")
    sum_columns = [
        ("rutenummer", "Rutenummer", 14),
        ("rutenavn", "Rutenavn", 30),
        ("vedlikeholdsansvarlig", "Vedlikeholdsansvarlig", 28),
        ("status", "Status", 12),
        ("errors", "Feil", 8),
        ("warnings", "Advarsler", 11),
        ("info", "Info", 8),
    ]
    _write_header(ws2, sum_columns)

    srow = 2
    for result in sorted(results, key=lambda r: r.rutenummer or ""):
        rutenavn, vedl = route_meta_by_rn.get(result.rutenummer, (None, None))
        status = result.get_status()
        vals = [
            result.rutenummer,
            rutenavn or "",
            vedl or "",
            status,
            len(result.errors),
            len(result.warnings),
            len(result.info),
        ]
        fill_key = {"ERROR": "ERROR", "WARNING": "WARNING", "OK": "OK"}.get(status, "INFO")
        fill = SEVERITY_FILL.get(fill_key)
        for ci, v in enumerate(vals, start=1):
            cell = ws2.cell(row=srow, column=ci, value=v)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill
        srow += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
