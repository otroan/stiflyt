"""Excel manufacturing list for the signs_app.

One row per panel. Distances use the corrected/rounded display value.
Layout chosen so a sign manufacturer can read it directly; the same sheet
is re-importable for color/distance overrides (see :func:`parse_uploaded_workbook`).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .sign_candidates import get_sign_candidates_for_area


HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
NUM_RIGHT = Alignment(horizontal="right")

# Manufacturer-facing columns. One row per panel; site-level fields
# (baksidetekst, sendes_til) repeat across the panels of a site.
COLUMNS = [
    ("antall", "Antall", 7),
    ("tekst_paa_skiltet", "Tekst på skiltet", 26),
    ("km", "Km", 7),
    ("pilretning", "Pilretning", 12),
    ("baksidetekst", "Baksidetekst", 36),
    ("skiltfarge", "Skiltfarge", 11),
    ("ruter", "Ruter", 14),
    ("sendes_til", "Sendes til (navn, adresse)", 30),
]


def _selection_key(sign_site_id: int, panel: Dict[str, Any]) -> str:
    """3-part selection key matching the frontend: <sign_site_id>:<anchor>:<first_link>.

    Parallel-path panels share (sign_site_id, destination_anchor_node_id) but
    differ on first_link_id; we have to key on all three or the filter will
    keep both siblings when the user only ticked one.
    """
    aid = panel.get("destination_anchor_node_id")
    fl = panel.get("first_link_id")
    return f"{sign_site_id}:{aid if aid is not None else ''}:{fl if fl is not None else ''}"


def _format_km(v: Any) -> str:
    if v is None:
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ""
    return f"{f:.1f}" if f < 10 else str(int(round(f)))


def build_manufacturing_workbook(
    conn,
    area_code: str,
    selection: Any = None,
) -> bytes:
    """Render the manufacturing list for one area as an .xlsx file.

    `selection`, when provided, is an iterable of ``"<sign_site_id>:<destination_anchor_node_id>"``
    strings (matches what the frontend builds). Only panels whose key matches
    are emitted. None / empty selection = export everything.
    """
    data = get_sign_candidates_for_area(conn, area_code)
    selection_set = None
    if selection:
        selection_set = {str(s) for s in selection if s}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Skilt {area_code}"

    for col_idx, (_key, header, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    row = 2
    for site in data.get("sites", []):
        panels: List[Dict[str, Any]] = site.get("panels") or []
        if selection_set is not None:
            sid = site.get("sign_site_id")
            if sid is None:
                # Skip unaccepted candidates entirely when filtering by selection
                continue
            panels = [
                p for p in panels
                if _selection_key(sid, p) in selection_set
            ]
            if not panels:
                continue
        if not panels:
            _write_row(ws, row, site, None)
            row += 1
            continue
        for panel in panels:
            _write_row(ws, row, site, panel)
            row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_row(ws, row: int, site: Dict[str, Any], panel: Dict[str, Any] | None) -> None:
    send_to_name = (site.get("send_to_name") or "").strip()
    send_to_address = (site.get("send_to_address") or "").strip()
    sendes_til = "\n".join(p for p in (send_to_name, send_to_address) if p)
    values = {
        "antall": 1 if panel else "",
        "tekst_paa_skiltet": (panel.get("destination_name") if panel else "") or "",
        "km": _format_km(panel.get("distance_km_displayed")) if panel else "",
        "pilretning": (panel.get("direction") if panel else "") or "",
        "baksidetekst": site.get("back_text") or "",
        "skiltfarge": (panel.get("color") if panel else "trehvit") or "trehvit",
        "ruter": ", ".join(panel.get("route_numbers") or []) if panel else ", ".join(site.get("route_numbers") or []),
        "sendes_til": sendes_til,
    }
    for col_idx, (key, _h, _w) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=row, column=col_idx, value=values[key])
        c.font = BODY_FONT
        if key in ("km", "antall"):
            c.alignment = NUM_RIGHT
        elif key in ("baksidetekst", "tekst_paa_skiltet", "sendes_til"):
            c.alignment = WRAP


def parse_uploaded_workbook(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Read an edited manufacturing workbook back; returns one dict per row.

    Used by the re-import endpoint to harvest color/distance overrides. Tolerates
    re-ordered columns by matching header names; ignores empty rows. No DB writes
    happen here — the caller decides how to apply the overrides.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    header_to_key = {h: key for key, h, _w in COLUMNS}
    key_index: Dict[str, int] = {}
    for idx, h in enumerate(headers):
        key = header_to_key.get(h)
        if key is not None:
            key_index[key] = idx
    out: List[Dict[str, Any]] = []
    for raw in rows[1:]:
        if not raw or all(v in (None, "") for v in raw):
            continue
        rec: Dict[str, Any] = {}
        for key, idx in key_index.items():
            rec[key] = raw[idx] if idx < len(raw) else None
        out.append(rec)
    return out
