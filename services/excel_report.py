"""Service for generating Excel reports for route owners."""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .matrikkel_owner_service import fetch_owners_for_matrikkelenheter, analyze_owner_fetch_errors
from .database import db_connection, get_teig_schema


def get_teig_data_source_info():
    """
    Get information about when teig data was last updated.
    Tries multiple methods to get update date, falls back to schema name.

    Returns:
        str: Data source information (e.g., "Teig data oppdatert: 2024-01-15" or schema name)
    """
    try:
        with db_connection() as conn:
            teig_schema = get_teig_schema(conn)

            with conn.cursor() as cur:
                # Method 1: Try to get last modified from pg_stat_user_tables
                try:
                    cur.execute("""
                        SELECT MAX(stat_last_modified) as last_modified
                        FROM pg_stat_user_tables
                        WHERE schemaname = %s;
                    """, (teig_schema,))
                    result = cur.fetchone()
                    if result and result[0]:
                        if isinstance(result[0], datetime):
                            date_str = result[0].strftime('%Y-%m-%d')
                        else:
                            date_str = str(result[0])[:10]  # Take first 10 chars for date
                        return f"Teig data oppdatert: {date_str}"
                except Exception:
                    pass

                # Method 2: Try to get from pg_stat_database or pg_class
                try:
                    cur.execute("""
                        SELECT MAX(reltuples::bigint) as row_count
                        FROM pg_class c
                        JOIN pg_namespace n ON n.oid = c.relnamespace
                        WHERE n.nspname = %s AND c.relkind = 'r';
                    """, (teig_schema,))
                    # This doesn't give us date, but we can use schema name as fallback
                except Exception:
                    pass

            # Fallback: use schema name (always 'stiflyt' now)
            return f"Teig schema: {teig_schema}"
    except Exception as e:
        # If we can't get info, return generic message
        print(f"Could not get teig data source info: {e}")
        return "Teig data: Ukjent oppdateringsdato"


def generate_owners_excel_from_data(matrikkelenhet_vector, metadata=None, title="Rapport"):
    """
    Generate Excel report from matrikkelenhet_vector data directly.

    Args:
        matrikkelenhet_vector: List of matrikkelenhet items with offset and length info
        metadata: Optional metadata dict with rutenummer, rutenavn, total_length_km, etc.
        title: Title for the report (used in filename and metadata)

    Returns:
        bytes: Excel file as bytes
    """
    if metadata is None:
        metadata = {}

    # Fetch owner information from Matrikkel API (if credentials are available)
    owner_results = fetch_owners_for_matrikkelenheter(matrikkelenhet_vector)

    # Check for errors - if any errors found, do not generate Excel report
    error_analysis = analyze_owner_fetch_errors(owner_results)
    if error_analysis['has_errors']:
        raise ValueError(error_analysis['error_summary'])

    # Create a mapping from matrikkelenhet identifier to owner info for quick lookup
    owner_info_map = {}
    for item, owner_info, error in owner_results:
        key = (
            item.get('kommunenummer'),
            item.get('gardsnummer'),
            item.get('bruksnummer'),
            item.get('festenummer'),
            item.get('matrikkelenhet', '')
        )
        owner_info_map[key] = owner_info if owner_info else ""

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Eiere"

    # Define header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define column headers
    headers = [
        "Offset (m)",
        "Offset (km)",
        "Lengde (m)",
        "Lengde (km)",
        "Matrikkelenhet",
        "Bruksnavn",
        "Kontaktinformasjon"
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Get current date/time for report generation
    generated_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Get teig data source information
    data_source_info = get_teig_data_source_info()

    # Write metadata rows
    metadata_row = 2
    if metadata:
        rutenavn = metadata.get('rutenavn', '')
        total_length = metadata.get('total_length_km', 0)
        rutenummer = metadata.get('rutenummer', title)

        # First row: Title/route info
        metadata_text = f"Utvalg: {rutenummer}" if rutenummer else f"Utvalg: {title}"
        if rutenavn:
            metadata_text += f" - {rutenavn}"
        if total_length > 0:
            metadata_text += f" | Total lengde: {total_length:.2f} km"
        ws.cell(row=metadata_row, column=1, value=metadata_text)
        ws.merge_cells(f'A{metadata_row}:G{metadata_row}')
        ws.cell(row=metadata_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[metadata_row].height = 20
        metadata_row += 1

    # Second row: Generation date
    ws.cell(row=metadata_row, column=1, value=f"Rapport generert: {generated_date}")
    ws.merge_cells(f'A{metadata_row}:G{metadata_row}')
    ws.cell(row=metadata_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[metadata_row].height = 20
    metadata_row += 1

    # Third row: Data source
    ws.cell(row=metadata_row, column=1, value=data_source_info)
    ws.merge_cells(f'A{metadata_row}:G{metadata_row}')
    ws.cell(row=metadata_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[metadata_row].height = 20
    metadata_row += 1

    # Write data rows
    start_row = metadata_row
    for idx, item in enumerate(matrikkelenhet_vector, start=start_row):
        ws.cell(row=idx, column=1, value=item.get('offset_meters', 0))
        ws.cell(row=idx, column=2, value=round(item.get('offset_km', 0), 3))
        ws.cell(row=idx, column=3, value=item.get('length_meters', 0))
        ws.cell(row=idx, column=4, value=round(item.get('length_km', 0), 3))
        ws.cell(row=idx, column=5, value=item.get('matrikkelenhet', ''))
        ws.cell(row=idx, column=6, value=item.get('bruksnavn', ''))
        key = (
            item.get('kommunenummer'),
            item.get('gardsnummer'),
            item.get('bruksnummer'),
            item.get('festenummer'),
            item.get('matrikkelenhet', '')
        )
        kontaktinformasjon = owner_info_map.get(key, "")
        # Split by semicolon and put each owner on a new line
        if kontaktinformasjon and ';' in kontaktinformasjon:
            owners = [owner.strip() for owner in kontaktinformasjon.split(';')]
            kontaktinformasjon = '\n'.join(owners)
        kontakt_cell = ws.cell(row=idx, column=7, value=kontaktinformasjon)
        kontakt_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Format columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 40

    # Format numeric columns
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = '#,##0'
        ws.cell(row=row, column=2).number_format = '#,##0.000'
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).number_format = '#,##0.000'

    # Set header row height
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = f'A{start_row}'

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()

